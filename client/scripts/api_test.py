"""
Gestor App - Script de Prueba API
==================================
Este script demuestra cómo usar la API para:
1. Obtener movimientos existentes
2. Crear nuevos movimientos (individual y bulk)
3. Importar desde Excel

REQUISITOS:
- pip install requests openpyxl

USO:
1. Genera un token en la app: Configuración > API & Desarrolladores
2. Reemplaza API_TOKEN y BASE_URL abajo
3. Ejecuta: python api_test.py
"""

import requests
from datetime import datetime
from typing import Optional

# ============================================================
# CONFIGURACIÓN - EDITAR ESTOS VALORES
# ============================================================
BASE_URL = "https://TU_APP.vercel.app"  # Cambiar por tu URL de Vercel
API_TOKEN = "sk_live_TU_TOKEN_AQUI"     # Pegar tu token aquí

# ============================================================
# Cliente API
# ============================================================
class GestorAPI:
    def __init__(self, base_url: str, token: str):
        self.base_url = base_url.rstrip('/')
        self.headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
    
    def get_movements(
        self, 
        limit: int = 50, 
        offset: int = 0,
        kind: Optional[str] = None,
        from_date: Optional[str] = None,
        to_date: Optional[str] = None,
        search: Optional[str] = None
    ) -> dict:
        """Obtener lista de movimientos con filtros opcionales"""
        params = {"limit": limit, "offset": offset}
        if kind:
            params["kind"] = kind
        if from_date:
            params["from"] = from_date
        if to_date:
            params["to"] = to_date
        if search:
            params["search"] = search
            
        response = requests.get(
            f"{self.base_url}/api/v1/movements",
            headers=self.headers,
            params=params
        )
        response.raise_for_status()
        return response.json()
    
    def create_movement(
        self,
        kind: str,  # 'income', 'expense', 'investment'
        amount: float,
        date: str,  # 'YYYY-MM-DD'
        account_id: str,
        description: Optional[str] = None,
        category_id: Optional[str] = None,
        provider: Optional[str] = None,
        payment_method: Optional[str] = None,
        tax_rate: Optional[float] = None,
        paid_by_external: Optional[str] = None,
        create_debt: bool = False
    ) -> dict:
        """Crear un nuevo movimiento"""
        data = {
            "kind": kind,
            "amount": amount,
            "date": date,
            "account_id": account_id
        }
        
        if description:
            data["description"] = description
        if category_id:
            data["category_id"] = category_id
        if provider:
            data["provider"] = provider
        if payment_method:
            data["payment_method"] = payment_method
        if tax_rate is not None:
            data["tax_rate"] = tax_rate
        if paid_by_external:
            data["paid_by_external"] = paid_by_external
            data["create_debt"] = create_debt
            
        response = requests.post(
            f"{self.base_url}/api/v1/movements",
            headers=self.headers,
            json=data
        )
        response.raise_for_status()
        return response.json()
    
    def create_movements_bulk(self, movements: list) -> dict:
        """Crear múltiples movimientos de una vez"""
        response = requests.post(
            f"{self.base_url}/api/v1/movements",
            headers=self.headers,
            json=movements
        )
        response.raise_for_status()
        return response.json()
    
    def health_check(self) -> dict:
        """Verificar estado de la API"""
        response = requests.get(f"{self.base_url}/api/v1")
        return response.json()


# ============================================================
# Funciones de Ejemplo
# ============================================================
def ejemplo_listar_movimientos(api: GestorAPI):
    """Ejemplo: Obtener los últimos 10 gastos"""
    print("\n📋 Obteniendo últimos 10 gastos...")
    result = api.get_movements(limit=10, kind="expense")
    
    print(f"   Total en base de datos: {result['total']}")
    print(f"   Devueltos: {result['count']}")
    
    for mov in result['data'][:5]:
        print(f"   - {mov['date']}: {mov['description']} -> {mov['amount']}€")


def ejemplo_crear_gasto(api: GestorAPI, account_id: str):
    """Ejemplo: Crear un gasto simple"""
    print("\n💸 Creando gasto de ejemplo...")
    
    result = api.create_movement(
        kind="expense",
        amount=25.50,
        date=datetime.now().strftime("%Y-%m-%d"),
        account_id=account_id,
        description="Prueba desde Python API",
        provider="Script Test"
    )
    
    if result.get('success'):
        print(f"   ✅ Creado! ID: {result['data'][0]['id']}")
    else:
        print(f"   ❌ Error: {result}")


def ejemplo_importar_desde_excel(api: GestorAPI, account_id: str, excel_path: str):
    """
    Ejemplo: Importar gastos desde un archivo Excel
    El Excel debe tener columnas: Fecha, Concepto, Importe, Categoría (opcional)
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ Necesitas instalar openpyxl: pip install openpyxl")
        return
    
    print(f"\n📊 Importando desde {excel_path}...")
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    movements = []
    for row in ws.iter_rows(min_row=2):  # Saltar cabecera
        fecha = row[0].value
        concepto = row[1].value
        importe = row[2].value
        
        if fecha and importe:
            movements.append({
                "kind": "expense",
                "amount": float(importe),
                "date": fecha.strftime("%Y-%m-%d") if hasattr(fecha, 'strftime') else str(fecha),
                "account_id": account_id,
                "description": concepto or ""
            })
    
    if movements:
        result = api.create_movements_bulk(movements)
        print(f"   ✅ Importados: {result.get('created', 0)} movimientos")
    else:
        print("   ⚠️ No se encontraron movimientos válidos")


# ============================================================
# Main
# ============================================================
if __name__ == "__main__":
    print("=" * 50)
    print("🚀 Gestor App - Test de API")
    print("=" * 50)
    
    # Verificar configuración
    if "TU_APP" in BASE_URL or "TU_TOKEN" in API_TOKEN:
        print("\n⚠️  CONFIGURA BASE_URL y API_TOKEN antes de ejecutar")
        print("   1. Cambia BASE_URL por tu URL de Vercel")
        print("   2. Genera un token en: Configuración > API & Desarrolladores")
        print("   3. Pega el token en API_TOKEN")
        exit(1)
    
    api = GestorAPI(BASE_URL, API_TOKEN)
    
    # Health check
    print("\n🏥 Verificando API...")
    try:
        health = api.health_check()
        print(f"   Status: {health.get('status', 'unknown')}")
        print(f"   Version: {health.get('version', 'unknown')}")
    except Exception as e:
        print(f"   ❌ Error conectando: {e}")
        exit(1)
    
    # Listar movimientos
    ejemplo_listar_movimientos(api)
    
    # Para crear movimientos, necesitas un account_id válido
    # Descomenta y reemplaza con tu ID:
    # ejemplo_crear_gasto(api, "tu-account-uuid-aqui")
    
    # Para importar desde Excel:
    # ejemplo_importar_desde_excel(api, "tu-account-uuid", "gastos.xlsx")
    
    print("\n✅ Test completado!")
